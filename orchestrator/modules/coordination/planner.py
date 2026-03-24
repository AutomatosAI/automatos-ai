"""
Mission Planner — PRD-82A + 82B
=================================

Goal decomposition into a task DAG with structural validation.

The planner:
  1. Tries template matching first (82B) — keyword-based, no LLM call
  2. Falls back to LLM decomposition if no template matches
  3. Validates the plan (DAG acyclic, agents exist, task count in bounds)
  4. Retries LLM up to 3 times on validation failure
  5. Returns a DecompositionResult

Source: PRD-82A Section 12 (US-011), PRD-82B US-001/US-002
"""

import asyncio
import json
import logging
import re
from dataclasses import dataclass, field
from typing import Any, Dict, List, Optional, Sequence
from uuid import UUID, uuid4

from config import COMPLEXITY_TOKEN_BUDGET
from core.llm import create_llm_manager
from core.models.core import Agent
from core.models.orchestration_enums import ComplexityTier, TaskType
from modules.coordination.templates import match_template, render_template
from services.orchestration_deps import (
    CyclicDependencyError,
    DependencyResolver,
    InvalidDependencyError,
)

logger = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Attachment helpers
# ---------------------------------------------------------------------------


def _extract_office_text(raw: bytes, filename_lower: str) -> str:
    """Extract text from Office documents (.docx, .xlsx)."""
    import io
    if filename_lower.endswith(".docx"):
        from docx import Document as DocxDocument
        doc = DocxDocument(io.BytesIO(raw))
        return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())
    elif filename_lower.endswith(".xlsx") or filename_lower.endswith(".xls"):
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
        lines = []
        for ws in wb.worksheets:
            lines.append(f"## Sheet: {ws.title}")
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                if any(cells):
                    lines.append(" | ".join(cells))
        wb.close()
        return "\n".join(lines)
    else:
        return f"[Office file: unsupported format]"


async def _fetch_attachment_contents(
    attachments: List[Dict[str, Any]],
) -> List[Dict[str, str]]:
    """Fetch text content from S3 attachment keys.

    Returns a list of {filename, content} dicts. Binary files are skipped
    with a placeholder note.
    """
    if not attachments:
        return []

    try:
        import boto3
        from botocore.config import Config as BotoConfig
        from core.config import Config

        boto_cfg = BotoConfig(region_name=Config.AWS_REGION)
        s3 = boto3.client(
            "s3",
            aws_access_key_id=Config.AWS_ACCESS_KEY_ID,
            aws_secret_access_key=Config.AWS_SECRET_ACCESS_KEY,
            config=boto_cfg,
        )
    except Exception:
        logger.warning("Cannot initialize S3 client for attachments — skipping")
        return []

    loop = asyncio.get_running_loop()
    results: List[Dict[str, str]] = []

    for att in attachments:
        # New format: document_id (from DocumentManager pipeline)
        doc_id = att.get("document_id")
        if doc_id:
            try:
                from core.database.database import SessionLocal
                from core.models.core import Document
                db = SessionLocal()
                try:
                    doc = db.query(Document).filter(Document.id == doc_id).first()
                    if doc and doc.file_path:
                        s3_path = doc.file_path
                        if s3_path.startswith("s3://"):
                            parts = s3_path[5:].split("/", 1)
                            bucket, key = parts[0], parts[1] if len(parts) > 1 else ""
                        else:
                            bucket, key = Config.S3_DOCUMENTS_BUCKET, s3_path
                        response = await loop.run_in_executor(
                            None,
                            lambda b=bucket, k=key: s3.get_object(Bucket=b, Key=k),
                        )
                        raw = response["Body"].read()
                        fname = doc.filename or att.get("filename", f"doc_{doc_id}")
                        if fname.lower().endswith(".pdf"):
                            try:
                                import fitz
                                pdf_doc = fitz.open(stream=raw, filetype="pdf")
                                content = "\n\n".join(p.get_text() for p in pdf_doc)
                                pdf_doc.close()
                            except ImportError:
                                content = raw.decode("utf-8", errors="replace")
                        else:
                            content = raw.decode("utf-8", errors="replace")
                        results.append({"filename": fname, "content": content})
                        logger.info("Fetched attachment doc_id=%s: %s (%d chars)", doc_id, fname, len(content))
                    else:
                        logger.warning("Attachment doc_id=%s not found or no file_path", doc_id)
                finally:
                    db.close()
            except Exception as exc:
                logger.warning("Failed to fetch attachment doc_id=%s: %s", doc_id, exc)
            continue

        # Legacy format: S3 key directly
        s3_key = att.get("key", "")
        filename = att.get("filename", "unknown")
        content_type = att.get("content_type", "")

        # Validate S3 key — must be under workspaces/ prefix
        if s3_key and not s3_key.startswith("workspaces/"):
            logger.warning("Skipping attachment with disallowed S3 key prefix: %s", s3_key[:100])
            results.append({
                "filename": filename,
                "content": f"[Attachment skipped — invalid storage path]",
            })
            continue

        # Determine extractable file types
        text_types = {
            "text/", "application/json", "application/pdf",
            "text/csv", "text/markdown", "text/plain",
        }
        office_extensions = {".docx", ".xlsx", ".doc", ".xls"}
        is_text = any(content_type.startswith(t) for t in text_types)
        fname_lower = filename.lower()
        is_office = any(fname_lower.endswith(ext) for ext in office_extensions)
        is_pdf = "pdf" in fname_lower or content_type == "application/pdf"

        if not is_text and not is_pdf and not is_office:
            results.append({
                "filename": filename,
                "content": f"[Binary file: {filename} ({content_type}) — content not extracted]",
            })
            continue

        try:
            response = await loop.run_in_executor(
                None,
                lambda k=s3_key: s3.get_object(
                    Bucket=Config.S3_DOCUMENTS_BUCKET, Key=k,
                ),
            )
            raw = response["Body"].read()

            # PDF extraction
            if is_pdf:
                try:
                    import fitz  # PyMuPDF
                    doc = fitz.open(stream=raw, filetype="pdf")
                    text_parts = [page.get_text() for page in doc]
                    doc.close()
                    content = "\n\n".join(text_parts)
                except ImportError:
                    content = "[PDF file — PyMuPDF not available for text extraction]"
            # Office document extraction
            elif is_office:
                try:
                    content = _extract_office_text(raw, fname_lower)
                except Exception as exc:
                    logger.warning("Office extraction failed for %s: %s", filename, exc)
                    content = f"[Office file: {filename} — text extraction failed]"
            else:
                content = raw.decode("utf-8", errors="replace")

            results.append({"filename": filename, "content": content})
            logger.info("Fetched attachment: %s (%d chars)", filename, len(content))

        except Exception as exc:
            logger.warning("Failed to fetch attachment %s: %s", s3_key, exc)
            results.append({
                "filename": filename,
                "content": f"[Failed to fetch: {exc}]",
            })

    return results


# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

MIN_TASKS = 3
MAX_TASKS = 20
MAX_PLAN_RETRIES = 3
TOKENS_PER_TASK_ESTIMATE = 2000  # legacy fallback


def _estimate_token_budget(tasks: List["PlannedTask"]) -> int:
    """Sum token budgets based on each task's complexity tier."""
    return sum(
        COMPLEXITY_TOKEN_BUDGET.get(t.complexity, TOKENS_PER_TASK_ESTIMATE)
        for t in tasks
    )

# Deliverable keywords for complexity scoring
_DELIVERABLE_KEYWORDS = frozenset({
    "report", "paper", "app", "application", "analysis",
    "presentation", "dashboard", "pipeline", "system",
})

# Domain topic clusters for breadth estimation.
# Multi-word terms use substring matching; single-word terms use word-boundary matching.
_DOMAIN_CLUSTERS: List[frozenset[str]] = [
    frozenset({"ai", "machine learning", "ml", "deep learning", "neural", "llm", "nlp", "gpt", "coordination"}),
    frozenset({"web", "frontend", "backend", "api", "rest", "graphql", "react", "html", "css"}),
    frozenset({"data", "database", "sql", "analytics", "etl", "warehouse", "visualization"}),
    frozenset({"security", "authentication", "encryption", "oauth", "compliance"}),
    frozenset({"cloud", "aws", "azure", "gcp", "kubernetes", "docker", "devops", "ci/cd"}),
    frozenset({"mobile", "ios", "android", "swift", "kotlin", "flutter"}),
    frozenset({"business", "marketing", "finance", "strategy", "revenue", "growth"}),
    frozenset({"research", "experiment", "prior art", "literature", "survey", "implications"}),
    frozenset({"design", "ux", "figma", "wireframe", "prototype"}),
    frozenset({"testing", "qa", "automation", "performance", "load testing"}),
]


# ---------------------------------------------------------------------------
# Complexity detection (PRD-82C US-004)
# ---------------------------------------------------------------------------


def _count_deliverables(goal: str) -> int:
    """Count how many deliverable keywords appear in the goal."""
    goal_lower = goal.lower()
    return sum(1 for kw in _DELIVERABLE_KEYWORDS if kw in goal_lower)


def _estimate_domains(goal: str) -> int:
    """Count how many distinct topic clusters the goal spans."""
    goal_lower = goal.lower()
    goal_words = set(re.findall(r"[a-z0-9/\-]+", goal_lower))

    def _cluster_matches(cluster: frozenset[str]) -> bool:
        for term in cluster:
            if " " in term:
                # Multi-word: substring match
                if term in goal_lower:
                    return True
            else:
                # Single-word: word-boundary match
                if term in goal_words:
                    return True
        return False

    return sum(1 for cluster in _DOMAIN_CLUSTERS if _cluster_matches(cluster))


def _detect_complexity(
    goal: str,
    attachments: Optional[List[Any]] = None,
) -> ComplexityTier:
    """
    Score goal complexity and return the appropriate tier.

    Signals:
      - word_count > 50 → +1
      - deliverable_count >= 1 → +1, >= 3 → +1 (bonus)
      - domain_breadth >= 2 → +1, >= 4 → +1 (bonus)
      - attachment_count > 0 → +1

    Score >= 3 → COMPLEX, >= 1 → MODERATE, else SIMPLE.
    """
    score = 0

    if len(goal.split()) > 50:
        score += 1

    deliverable_count = _count_deliverables(goal)
    if deliverable_count >= 1:
        score += 1
    if deliverable_count >= 3:
        score += 1

    domain_count = _estimate_domains(goal)
    if domain_count >= 2:
        score += 1
    if domain_count >= 4:
        score += 1

    if attachments and len(attachments) > 0:
        score += 1

    if score >= 3:
        return ComplexityTier.COMPLEX
    elif score >= 1:
        return ComplexityTier.MODERATE
    else:
        return ComplexityTier.SIMPLE


def _complexity_to_max_concurrent(tier: ComplexityTier) -> int:
    """Map complexity tier to max_concurrent value."""
    if tier == ComplexityTier.COMPLEX:
        return 3
    elif tier == ComplexityTier.MODERATE:
        return 2
    return 1


# ---------------------------------------------------------------------------
# Result dataclasses
# ---------------------------------------------------------------------------


@dataclass(frozen=True)
class PlannedTask:
    """A single task in the decomposition result."""

    temp_id: str
    title: str
    description: str
    agent_role: str
    sequence_number: int
    task_type: str
    verification_criteria: List[Dict[str, Any]]
    required_tools: List[str]
    dependencies: List[str]  # temp_ids of upstream tasks
    complexity: str = "moderate"
    parallel_group: Optional[str] = None


@dataclass(frozen=True)
class PlannedDependency:
    """A dependency edge between two planned tasks."""

    from_task_temp_id: str
    to_task_temp_id: str


@dataclass(frozen=True)
class DecompositionResult:
    """Immutable result of goal decomposition."""

    tasks: List[PlannedTask]
    dependencies: List[PlannedDependency]
    token_estimate: int
    template_used: Optional[str] = None
    max_concurrent: int = 1


# ---------------------------------------------------------------------------
# Exceptions
# ---------------------------------------------------------------------------


class PlanValidationError(Exception):
    """Raised when the plan fails structural validation after all retries."""

    def __init__(self, errors: List[str]):
        self.errors = errors
        super().__init__(f"Plan validation failed: {'; '.join(errors)}")


# ---------------------------------------------------------------------------
# MissionPlanner
# ---------------------------------------------------------------------------


class MissionPlanner:
    """
    Decomposes a goal into a validated task DAG using an LLM call.

    Stateless — all data comes from arguments. LLM manager is created
    once and reused for retries within the same decompose() call.
    """

    @staticmethod
    async def replan(
        goal: str,
        workspace_id: UUID,
        agents: Sequence[Agent],
        completed_outputs: List[Dict[str, Any]],
        failed_task_title: str,
        failed_task_reason: str,
        user_notes: Optional[str] = None,
        config: Optional[Dict[str, Any]] = None,
    ) -> DecompositionResult:
        """
        Replan a failed mission — generate replacement tasks for the failed
        subtree while preserving completed/verified tasks.

        Args:
            goal: Original mission goal.
            workspace_id: Owning workspace UUID.
            agents: Available roster agents.
            completed_outputs: List of dicts with keys: task_id, title, output
                               (summaries of already-completed tasks).
            failed_task_title: Title of the task that failed.
            failed_task_reason: Reason the task failed.
            user_notes: Optional user guidance for replanning.
            config: Optional overrides.

        Returns:
            DecompositionResult with replacement tasks and dependencies.

        Raises:
            PlanValidationError: if all retry attempts fail structural validation.
        """
        llm = create_llm_manager(service_name="orchestrator")
        agent_roster = _render_agent_roster(agents)
        last_errors: List[str] = []

        for attempt in range(1, MAX_PLAN_RETRIES + 1):
            logger.info(
                "MissionPlanner.replan: attempt %d/%d for goal='%s' workspace=%s",
                attempt,
                MAX_PLAN_RETRIES,
                goal[:80],
                workspace_id,
            )

            prompt = _build_replan_prompt(
                goal=goal,
                agent_roster=agent_roster,
                completed_outputs=completed_outputs,
                failed_task_title=failed_task_title,
                failed_task_reason=failed_task_reason,
                user_notes=user_notes,
                validation_errors=last_errors if attempt > 1 else None,
            )

            messages = [
                {"role": "system", "content": _REPLAN_SYSTEM_PROMPT},
                {"role": "user", "content": prompt},
            ]

            try:
                response = await llm.generate_response(messages)
            except Exception:
                logger.error(
                    "MissionPlanner.replan: LLM call failed on attempt %d",
                    attempt,
                    exc_info=True,
                )
                last_errors = ["LLM call failed — retrying"]
                continue

            raw = _extract_json(response.content)
            if raw is None:
                last_errors = [
                    "LLM response did not contain valid JSON. "
                    "Ensure your response is a single JSON object."
                ]
                logger.warning(
                    "MissionPlanner.replan: no JSON in LLM response on attempt %d",
                    attempt,
                )
                continue

            parse_errors: List[str] = []
            tasks, deps = _parse_plan(raw, parse_errors)
            if parse_errors:
                last_errors = parse_errors
                logger.warning(
                    "MissionPlanner.replan: parse errors on attempt %d: %s",
                    attempt,
                    parse_errors,
                )
                continue

            tasks, deps = _ensure_synthesis_tasks(tasks, deps)
            validation_errors = _validate_plan(tasks, deps, agents)
            if validation_errors:
                last_errors = validation_errors
                logger.warning(
                    "MissionPlanner.replan: validation errors on attempt %d: %s",
                    attempt,
                    validation_errors,
                )
                continue

            token_estimate = _estimate_token_budget(tasks)
            # Compute max_concurrent same as decompose()
            max_concurrent = _complexity_to_max_concurrent(
                _detect_complexity(goal, None)
            )
            logger.info(
                "MissionPlanner.replan: generated %d replacement tasks (attempt %d, max_concurrent=%d)",
                len(tasks),
                attempt,
                max_concurrent,
            )
            return DecompositionResult(
                tasks=tasks,
                dependencies=deps,
                token_estimate=token_estimate,
                max_concurrent=max_concurrent,
            )

        raise PlanValidationError(last_errors)

    @staticmethod
    async def decompose(
        goal: str,
        workspace_id: UUID,
        agents: Sequence[Agent],
        config: Optional[Dict[str, Any]] = None,
    ) -> DecompositionResult:
        """
        Decompose *goal* into a task DAG validated against available *agents*.

        Args:
            goal: Natural-language goal string from the user.
            workspace_id: Owning workspace UUID.
            agents: Available roster agents for this workspace.
            config: Optional overrides (unused in v1, reserved for 82B).

        Returns:
            DecompositionResult with tasks, dependencies, and token estimate.

        Raises:
            PlanValidationError: if all retry attempts fail structural validation.
        """
        # --- Complexity detection (82C US-004) ---
        raw_attachments_list = (config or {}).get("attachments")
        complexity = _detect_complexity(goal, raw_attachments_list)
        max_concurrent = _complexity_to_max_concurrent(complexity)
        logger.info(
            "MissionPlanner: complexity=%s max_concurrent=%d for goal='%s'",
            complexity.value,
            max_concurrent,
            goal[:80],
        )

        # --- Template matching (82B US-002) — try before LLM ---
        template = match_template(goal)
        if template is not None:
            logger.info(
                "MissionPlanner: template=%s matched for goal='%s'",
                template.id,
                goal[:80],
            )
            raw_tasks = render_template(template, goal)
            parse_errors: List[str] = []
            tasks, deps = _parse_plan({"tasks": raw_tasks}, parse_errors)
            if not parse_errors:
                tasks, deps = _ensure_synthesis_tasks(tasks, deps)
                validation_errors = _validate_plan(tasks, deps, agents)
                if not validation_errors:
                    token_estimate = _estimate_token_budget(tasks)
                    logger.info(
                        "MissionPlanner: template=%s produced %d tasks",
                        template.id,
                        len(tasks),
                    )
                    return DecompositionResult(
                        tasks=tasks,
                        dependencies=deps,
                        token_estimate=token_estimate,
                        template_used=template.id,
                        max_concurrent=max_concurrent,
                    )
                else:
                    logger.warning(
                        "MissionPlanner: template=%s failed validation: %s — falling through to LLM",
                        template.id,
                        validation_errors,
                    )
            else:
                logger.warning(
                    "MissionPlanner: template=%s failed parsing: %s — falling through to LLM",
                    template.id,
                    parse_errors,
                )
        else:
            logger.info(
                "MissionPlanner: no template match, using LLM decomposition"
            )

        # --- LLM decomposition fallback ---
        llm = create_llm_manager(service_name="orchestrator")
        agent_roster = _render_agent_roster(agents)
        last_errors: List[str] = []

        # Fetch attachment contents if present in config
        attachment_contents: Optional[List[Dict[str, str]]] = None
        raw_attachments = (config or {}).get("attachments")
        if raw_attachments and isinstance(raw_attachments, list):
            attachment_contents = await _fetch_attachment_contents(raw_attachments)
            logger.info(
                "MissionPlanner: fetched %d attachment(s) for context",
                len(attachment_contents),
            )

        for attempt in range(1, MAX_PLAN_RETRIES + 1):
            logger.info(
                "MissionPlanner.decompose: attempt %d/%d for goal='%s' workspace=%s",
                attempt,
                MAX_PLAN_RETRIES,
                goal[:80],
                workspace_id,
            )

            prompt = _build_decomposition_prompt(
                goal=goal,
                agent_roster=agent_roster,
                validation_errors=last_errors if attempt > 1 else None,
                attachment_contents=attachment_contents,
            )

            messages = [
                {"role": "system", "content": _SYSTEM_PROMPT},
                {"role": "user", "content": prompt},
            ]

            try:
                response = await llm.generate_response(messages)
            except Exception:
                logger.error(
                    "MissionPlanner: LLM call failed on attempt %d",
                    attempt,
                    exc_info=True,
                )
                last_errors = ["LLM call failed — retrying"]
                continue

            raw = _extract_json(response.content)
            if raw is None:
                last_errors = [
                    "LLM response did not contain valid JSON. "
                    "Ensure your response is a single JSON object."
                ]
                logger.warning(
                    "MissionPlanner: no JSON in LLM response on attempt %d",
                    attempt,
                )
                continue

            # Parse into PlannedTask list
            parse_errors: List[str] = []
            tasks, deps = _parse_plan(raw, parse_errors)
            if parse_errors:
                last_errors = parse_errors
                logger.warning(
                    "MissionPlanner: parse errors on attempt %d: %s",
                    attempt,
                    parse_errors,
                )
                continue

            # Auto-insert synthesis tasks for parallel convergence (82C US-008)
            tasks, deps = _ensure_synthesis_tasks(tasks, deps)

            # Structural validation
            validation_errors = _validate_plan(tasks, deps, agents)
            if validation_errors:
                last_errors = validation_errors
                logger.warning(
                    "MissionPlanner: validation errors on attempt %d: %s",
                    attempt,
                    validation_errors,
                )
                continue

            # Success
            token_estimate = _estimate_token_budget(tasks)
            logger.info(
                "MissionPlanner: decomposed goal into %d tasks (attempt %d)",
                len(tasks),
                attempt,
            )
            return DecompositionResult(
                tasks=tasks,
                dependencies=deps,
                token_estimate=token_estimate,
                max_concurrent=max_concurrent,
            )

        # All retries exhausted
        raise PlanValidationError(last_errors)


# ---------------------------------------------------------------------------
# Prompt construction
# ---------------------------------------------------------------------------

_SYSTEM_PROMPT = """\
You are a mission planner for an AI agent platform. Your job is to decompose \
a user's goal into a task DAG that can be executed by the available agents. \
Tasks may run in parallel when they are independent.

Rules:
- Each task must be atomic — completable by ONE agent in ONE execution.
- Each task should produce at most ~4000 words of output.
- Use dependencies to encode ordering. Tasks with no dependencies on each other \
can run in parallel.
- Assign a complexity tier to each task: "simple" (~1000 tokens), "moderate" \
(~4000 tokens), or "complex" (~8000 tokens).
- Group independent tasks that can run in parallel under the same parallel_group \
name (e.g. "research", "analysis"). Tasks in the same parallel_group MUST NOT \
depend on each other.
- After parallel groups converge, include a "synthesis" task (task_type="synthesis") \
that merges and integrates the outputs of the parallel tasks.
- Every task must specify an agent_role matching one of the available agents.
- The plan MUST contain between 3 and 20 tasks inclusive.
- Return ONLY a single JSON object (no markdown, no explanation).
"""

_VALID_TASK_TYPES = frozenset(t.value for t in TaskType)
_VALID_COMPLEXITIES = frozenset({"simple", "moderate", "complex", "synthesis"})


def _build_decomposition_prompt(
    *,
    goal: str,
    agent_roster: str,
    validation_errors: Optional[List[str]] = None,
    attachment_contents: Optional[List[Dict[str, str]]] = None,
) -> str:
    """Build the user prompt for goal decomposition."""
    parts = [
        f"## Goal\n<user_goal>\n{goal}\n</user_goal>\n",
    ]

    # Inject attachment content so the LLM has full context
    if attachment_contents:
        parts.append("## Attached Reference Documents\n")
        total_budget = 90_000  # Global budget for all attachments combined
        per_file_limit = 30_000
        accumulated = 0
        for att in attachment_contents:
            remaining = total_budget - accumulated
            if remaining <= 0:
                parts.append(
                    "\n[... further attachments skipped — total attachment budget exceeded ...]\n"
                )
                break
            filename = att.get("filename", "unknown")
            content = att.get("content", "")
            limit = min(len(content), per_file_limit, remaining)
            if len(content) > limit:
                content = content[:limit] + "\n\n[... truncated ...]"
            accumulated += len(content)
            parts.append(
                f"### {filename}\n<attachment>\n{content}\n</attachment>\n"
            )

    parts.append(f"## Available Agents\n{agent_roster}\n")
    parts.append(_OUTPUT_SCHEMA_INSTRUCTIONS)

    if validation_errors:
        error_text = "\n".join(f"- {e}" for e in validation_errors)
        parts.append(
            f"\n## Previous Attempt Failed Validation\n"
            f"Fix these errors in your next response:\n{error_text}\n"
        )

    return "\n".join(parts)


_OUTPUT_SCHEMA_INSTRUCTIONS = """\
## Required JSON Output

Return ONLY a JSON object with this exact structure:

```
{
  "tasks": [
    {
      "temp_id": "task_1",
      "title": "Short descriptive title",
      "description": "Detailed instructions for the agent",
      "agent_role": "researcher",
      "sequence_number": 1,
      "task_type": "llm_generation",
      "complexity": "moderate",
      "parallel_group": "research",
      "verification_criteria": [
        {"type": "min_length", "value": 200, "must_pass": true},
        {"type": "required_sections", "value": ["## Summary", "## Findings"], "must_pass": false}
      ],
      "required_tools": [],
      "dependencies": []
    },
    {
      "temp_id": "task_2",
      "title": "Research secondary sources",
      "description": "...",
      "agent_role": "researcher",
      "sequence_number": 1,
      "task_type": "llm_generation",
      "complexity": "moderate",
      "parallel_group": "research",
      "verification_criteria": [],
      "required_tools": [],
      "dependencies": []
    },
    {
      "temp_id": "task_3",
      "title": "Synthesize research findings",
      "description": "Merge and integrate all research outputs into a coherent whole",
      "agent_role": "writer",
      "sequence_number": 2,
      "task_type": "synthesis",
      "complexity": "moderate",
      "parallel_group": null,
      "verification_criteria": [],
      "required_tools": [],
      "dependencies": ["task_1", "task_2"]
    }
  ]
}
```

Valid task_type values: llm_generation, tool_execution, analysis, synthesis, review
Valid complexity values: simple, moderate, complex
Dependencies reference temp_id values of upstream tasks that must complete first.
parallel_group: string name grouping independent parallel tasks (null if sequential).
Tasks in the same parallel_group MUST NOT depend on each other.
After parallel groups converge, include a synthesis task that merges the outputs.
"""


_REPLAN_SYSTEM_PROMPT = """\
You are a mission replanner for an AI agent platform. A mission has partially \
completed but one or more tasks failed. Your job is to generate REPLACEMENT \
tasks that pick up where the mission left off, taking into account what has \
already been accomplished.

Rules:
- Do NOT duplicate work already completed by verified tasks.
- Generate only the tasks needed to replace the failed task and complete the goal.
- Each task must be atomic — completable by ONE agent in ONE execution.
- Each task should produce at most ~4000 words of output.
- Use dependencies to encode ordering. Independent tasks can run in parallel.
- Assign a complexity tier to each task: "simple", "moderate", or "complex".
- Group independent parallel tasks under the same parallel_group name. \
Tasks in the same parallel_group MUST NOT depend on each other.
- After parallel groups converge, include a "synthesis" task (task_type="synthesis") \
that merges the parallel outputs.
- Every task must specify an agent_role matching one of the available agents.
- The plan MUST contain between 3 and 20 tasks inclusive.
- Return ONLY a single JSON object (no markdown, no explanation).
"""


def _build_replan_prompt(
    *,
    goal: str,
    agent_roster: str,
    completed_outputs: List[Dict[str, Any]],
    failed_task_title: str,
    failed_task_reason: str,
    user_notes: Optional[str] = None,
    validation_errors: Optional[List[str]] = None,
) -> str:
    """Build the user prompt for replanning a failed mission."""
    completed_summary = ""
    if completed_outputs:
        lines = []
        for co in completed_outputs:
            title = co.get("title", "Unknown")
            output = co.get("output", "")
            excerpt = output[:300] if output else "(no output)"
            lines.append(f"- **{title}**: {excerpt}")
        completed_summary = "\n".join(lines)
    else:
        completed_summary = "(No tasks completed yet)"

    parts = [
        f"## Original Goal\n<user_goal>\n{goal}\n</user_goal>\n",
        f"## Completed Tasks (DO NOT REDO)\n{completed_summary}\n",
        f"## Failed Task\n- **Title**: {failed_task_title}\n- **Failure Reason**: {failed_task_reason}\n",
    ]

    if user_notes:
        parts.append(f"## User Guidance for Replan\n{user_notes}\n")

    parts.append(f"## Available Agents\n{agent_roster}\n")
    parts.append(_OUTPUT_SCHEMA_INSTRUCTIONS)

    if validation_errors:
        error_text = "\n".join(f"- {e}" for e in validation_errors)
        parts.append(
            f"\n## Previous Attempt Failed Validation\n"
            f"Fix these errors in your next response:\n{error_text}\n"
        )

    return "\n".join(parts)


def _render_agent_roster(agents: Sequence[Agent]) -> str:
    """Render available agents as a concise description for the planner prompt."""
    if not agents:
        return "(No agents available)\n"

    lines: List[str] = []
    for agent in agents:
        if agent.status != "active":
            continue

        model_id = ""
        if isinstance(agent.model_config, dict):
            model_id = agent.model_config.get("model_id", "")

        skills_text = ""
        try:
            if agent.skills:
                skill_names = [s.name for s in agent.skills if s.name]
                if skill_names:
                    skills_text = f" | Skills: {', '.join(skill_names)}"
        except Exception:
            pass  # Detached session — skills not loaded

        tags_text = ""
        if agent.tags and isinstance(agent.tags, list):
            tags_text = f" | Tags: {', '.join(str(t) for t in agent.tags)}"

        desc = (agent.description or "")[:120]
        lines.append(
            f"- **{agent.name}** (role: {agent.name.lower()})"
            f" — {desc}"
            f"{skills_text}{tags_text}"
            f"{f' | Model: {model_id}' if model_id else ''}"
        )

    return "\n".join(lines) if lines else "(No active agents)\n"


# ---------------------------------------------------------------------------
# JSON extraction
# ---------------------------------------------------------------------------


def _extract_json(content: str) -> Optional[Dict[str, Any]]:
    """
    Extract a JSON object from LLM response content.

    Handles markdown code blocks (```json ... ```) and raw JSON.
    """
    if not content:
        return None

    # Try to extract from markdown code block first
    block_match = re.search(r"```(?:json)?\s*\n?(.*?)```", content, re.DOTALL)
    text = block_match.group(1).strip() if block_match else content.strip()

    try:
        parsed = json.loads(text)
        if isinstance(parsed, dict):
            return parsed
    except json.JSONDecodeError:
        pass

    # Fallback: find first { ... } block
    brace_match = re.search(r"\{.*\}", content, re.DOTALL)
    if brace_match:
        try:
            parsed = json.loads(brace_match.group())
            if isinstance(parsed, dict):
                return parsed
        except json.JSONDecodeError:
            pass

    return None


# ---------------------------------------------------------------------------
# Plan parsing
# ---------------------------------------------------------------------------


def _parse_plan(
    raw: Dict[str, Any],
    errors: List[str],
) -> tuple:
    """
    Parse raw JSON dict into (List[PlannedTask], List[PlannedDependency]).

    Appends parse errors to *errors* list. Returns ([], []) on failure.
    """
    raw_tasks = raw.get("tasks")
    if not isinstance(raw_tasks, list) or not raw_tasks:
        errors.append("JSON must contain a non-empty 'tasks' array")
        return [], []

    tasks: List[PlannedTask] = []
    deps: List[PlannedDependency] = []
    seen_ids: set = set()

    for i, rt in enumerate(raw_tasks):
        if not isinstance(rt, dict):
            errors.append(f"tasks[{i}] is not an object")
            continue

        temp_id = str(rt.get("temp_id", f"task_{i + 1}"))
        if temp_id in seen_ids:
            errors.append(f"Duplicate temp_id: {temp_id}")
            continue
        seen_ids.add(temp_id)

        title = str(rt.get("title", "")).strip()
        if not title:
            errors.append(f"tasks[{i}] ({temp_id}): missing title")
            continue

        description = str(rt.get("description", "")).strip()
        agent_role = str(rt.get("agent_role", "")).strip()
        if not agent_role:
            errors.append(f"tasks[{i}] ({temp_id}): missing agent_role")
            continue

        sequence_number = rt.get("sequence_number", i + 1)
        if not isinstance(sequence_number, int):
            try:
                sequence_number = int(sequence_number)
            except (TypeError, ValueError):
                sequence_number = i + 1

        task_type = str(rt.get("task_type", TaskType.LLM_GENERATION.value))
        if task_type not in _VALID_TASK_TYPES:
            task_type = TaskType.LLM_GENERATION.value

        verification_criteria = rt.get("verification_criteria", [])
        if not isinstance(verification_criteria, list):
            verification_criteria = []

        required_tools = rt.get("required_tools", [])
        if not isinstance(required_tools, list):
            required_tools = []

        task_deps = rt.get("dependencies", [])
        if not isinstance(task_deps, list):
            task_deps = []

        # PRD-82C: complexity and parallel_group
        complexity = str(rt.get("complexity", "moderate")).lower()
        if complexity not in _VALID_COMPLEXITIES:
            complexity = "moderate"

        raw_pg = rt.get("parallel_group")
        parallel_group = str(raw_pg).strip() if raw_pg else None

        tasks.append(
            PlannedTask(
                temp_id=temp_id,
                title=title,
                description=description,
                agent_role=agent_role,
                sequence_number=sequence_number,
                task_type=task_type,
                verification_criteria=verification_criteria,
                required_tools=[str(t) for t in required_tools],
                dependencies=[str(d) for d in task_deps],
                complexity=complexity,
                parallel_group=parallel_group,
            )
        )

        for dep_id in task_deps:
            deps.append(
                PlannedDependency(
                    from_task_temp_id=str(dep_id),
                    to_task_temp_id=temp_id,
                )
            )

    return tasks, deps


# ---------------------------------------------------------------------------
# Synthesis auto-insertion (PRD-82C US-008)
# ---------------------------------------------------------------------------


def _ensure_synthesis_tasks(
    tasks: List[PlannedTask],
    deps: List[PlannedDependency],
) -> tuple:
    """
    Auto-insert synthesis tasks when parallel branches converge without one.

    For each parallel_group with 2+ tasks, check whether any downstream task
    depends on members of that group. If the downstream task is NOT a synthesis
    task, insert an auto-generated synthesis task between the group and the
    downstream consumer.

    Returns a new (tasks, deps) tuple — inputs are not mutated.
    """
    # Index parallel groups with 2+ members
    group_members: Dict[str, List[PlannedTask]] = {}
    for task in tasks:
        if task.parallel_group:
            group_members.setdefault(task.parallel_group, []).append(task)

    # Only consider groups that actually have parallelism
    parallel_groups = {
        name: members
        for name, members in group_members.items()
        if len(members) >= 2
    }

    if not parallel_groups:
        return tasks, deps

    # Check which groups already have explicit synthesis tasks downstream
    member_ids_by_group: Dict[str, frozenset] = {
        name: frozenset(t.temp_id for t in members)
        for name, members in parallel_groups.items()
    }

    # For each group, find downstream tasks that depend on ANY member
    deps_by_target: Dict[str, List[str]] = {}
    for dep in deps:
        deps_by_target.setdefault(dep.to_task_temp_id, []).append(
            dep.from_task_temp_id
        )

    task_by_id = {t.temp_id: t for t in tasks}

    # Track which groups already have a synthesis consumer
    groups_with_synthesis: set = set()
    for task in tasks:
        if task.task_type == TaskType.SYNTHESIS.value:
            task_upstream = set(deps_by_target.get(task.temp_id, []))
            for group_name, member_ids in member_ids_by_group.items():
                if task_upstream & member_ids:
                    groups_with_synthesis.add(group_name)

    groups_needing_synthesis = set(parallel_groups.keys()) - groups_with_synthesis
    if not groups_needing_synthesis:
        return tasks, deps

    # Build new task and dep lists (immutable approach)
    new_tasks = list(tasks)
    new_deps = list(deps)

    for group_name in sorted(groups_needing_synthesis):
        member_ids = member_ids_by_group[group_name]
        members = parallel_groups[group_name]

        # Max sequence in this group — synthesis goes right after
        max_group_seq = max(t.sequence_number for t in members)
        synth_seq = max_group_seq + 1

        synth_id = f"synth_{group_name}"
        # Ensure unique temp_id
        existing_ids = {t.temp_id for t in new_tasks}
        suffix = 0
        while synth_id in existing_ids:
            suffix += 1
            synth_id = f"synth_{group_name}_{suffix}"

        synth_task = PlannedTask(
            temp_id=synth_id,
            title=f"Synthesize {group_name} outputs",
            description=(
                f"Merge and synthesize the outputs from the '{group_name}' "
                f"parallel tasks into a unified, coherent result."
            ),
            agent_role="writer",
            sequence_number=synth_seq,
            task_type=TaskType.SYNTHESIS.value,
            verification_criteria=[],
            required_tools=[],
            dependencies=sorted(member_ids),
            complexity="synthesis",
            parallel_group=None,
        )

        new_tasks.append(synth_task)

        # Add deps: each parallel member → synth task
        for mid in sorted(member_ids):
            new_deps.append(
                PlannedDependency(
                    from_task_temp_id=mid,
                    to_task_temp_id=synth_id,
                )
            )

        # Re-point downstream tasks: any task that depended on a group member
        # should now depend on the synthesis task instead
        updated_deps: List[PlannedDependency] = []
        tasks_to_repoint: set = set()
        for dep in new_deps:
            # Skip deps we just added (member → synth)
            if dep.to_task_temp_id == synth_id:
                updated_deps.append(dep)
                continue
            # If a non-member task depends on a group member, repoint to synth
            if (
                dep.from_task_temp_id in member_ids
                and dep.to_task_temp_id not in member_ids
            ):
                if dep.to_task_temp_id not in tasks_to_repoint:
                    tasks_to_repoint.add(dep.to_task_temp_id)
                    updated_deps.append(
                        PlannedDependency(
                            from_task_temp_id=synth_id,
                            to_task_temp_id=dep.to_task_temp_id,
                        )
                    )
                # Drop the old direct dep (replaced by synth → downstream)
            else:
                updated_deps.append(dep)

        new_deps = updated_deps

        # Update repointed tasks' dependencies field for consistency
        updated_tasks: List[PlannedTask] = []
        for task in new_tasks:
            if task.temp_id in tasks_to_repoint:
                # Replace group member deps with synth dep
                old_deps_set = set(task.dependencies)
                new_task_deps = sorted(
                    (old_deps_set - member_ids) | {synth_id}
                )
                updated_tasks.append(
                    PlannedTask(
                        temp_id=task.temp_id,
                        title=task.title,
                        description=task.description,
                        agent_role=task.agent_role,
                        sequence_number=task.sequence_number,
                        task_type=task.task_type,
                        verification_criteria=task.verification_criteria,
                        required_tools=task.required_tools,
                        dependencies=new_task_deps,
                        complexity=task.complexity,
                        parallel_group=task.parallel_group,
                    )
                )
            else:
                updated_tasks.append(task)
        new_tasks = updated_tasks

    # Bump sequence numbers for tasks that come after inserted synthesis tasks
    # to maintain proper ordering
    logger.info(
        "_ensure_synthesis_tasks: auto-inserted %d synthesis task(s) for groups: %s",
        len(groups_needing_synthesis),
        sorted(groups_needing_synthesis),
    )

    return new_tasks, new_deps


# ---------------------------------------------------------------------------
# Plan validation
# ---------------------------------------------------------------------------


def _validate_plan(
    tasks: List[PlannedTask],
    deps: List[PlannedDependency],
    agents: Sequence[Agent],
) -> List[str]:
    """
    Structural validation of the decomposition plan.

    Returns a list of error strings (empty = valid).

    Checks:
      1. Task count within [MIN_TASKS, MAX_TASKS]
      2. All dependency references are valid temp_ids
      3. DAG is acyclic (via DependencyResolver)
      4. All agent_roles match at least one active agent
      5. No orphan tasks (every task reachable from a root)
    """
    errors: List[str] = []

    # 1. Task count
    if len(tasks) < MIN_TASKS:
        errors.append(
            f"Plan has {len(tasks)} tasks — minimum is {MIN_TASKS}"
        )
    elif len(tasks) > MAX_TASKS:
        errors.append(
            f"Plan has {len(tasks)} tasks — maximum is {MAX_TASKS}"
        )

    task_ids = {t.temp_id for t in tasks}

    # 2. Dependency references
    for dep in deps:
        if dep.from_task_temp_id not in task_ids:
            errors.append(
                f"Dependency references unknown task: {dep.from_task_temp_id}"
            )

    # If references are broken, skip DAG validation
    if errors:
        return errors

    # 3. DAG acyclicity — use UUID stand-ins for DependencyResolver
    temp_to_uuid = {t.temp_id: uuid4() for t in tasks}
    uuid_task_ids = [temp_to_uuid[t.temp_id] for t in tasks]

    # Build mock dependency objects with the fields DependencyResolver expects
    class _MockDep:
        def __init__(self, task_id: UUID, depends_on_task_id: UUID):
            self.task_id = task_id
            self.depends_on_task_id = depends_on_task_id

    mock_deps = [
        _MockDep(
            task_id=temp_to_uuid[d.to_task_temp_id],
            depends_on_task_id=temp_to_uuid[d.from_task_temp_id],
        )
        for d in deps
    ]

    try:
        DependencyResolver.validate_task_graph(uuid_task_ids, mock_deps)
    except CyclicDependencyError:
        errors.append("Task dependency graph contains a cycle")
    except InvalidDependencyError as exc:
        errors.append(str(exc))

    # 4. Agent role matching
    active_agent_names = set()
    for agent in agents:
        if agent.status == "active":
            active_agent_names.add(agent.name.lower())
            # Also match by skill names
            try:
                if agent.skills:
                    for s in agent.skills:
                        if s.name:
                            active_agent_names.add(s.name.lower())
            except Exception:
                pass
            # Also match by tags
            if agent.tags and isinstance(agent.tags, list):
                for t in agent.tags:
                    active_agent_names.add(str(t).lower())

    for task in tasks:
        role_lower = task.agent_role.lower()
        # Fuzzy: check if role matches any agent name, skill, or tag
        matched = any(
            role_lower == name or role_lower in name or name in role_lower
            for name in active_agent_names
        )
        if not matched:
            errors.append(
                f"Task '{task.title}' references agent_role '{task.agent_role}' "
                f"which does not match any active agent"
            )

    # 5. Parallel group cross-dependency check (PRD-82C)
    #    Tasks in the same parallel_group must NOT depend on each other.
    group_members: Dict[str, set] = {}
    for task in tasks:
        if task.parallel_group:
            group_members.setdefault(task.parallel_group, set()).add(task.temp_id)

    dep_set = {(d.from_task_temp_id, d.to_task_temp_id) for d in deps}
    for group_name, members in group_members.items():
        for dep_from, dep_to in dep_set:
            if dep_from in members and dep_to in members:
                errors.append(
                    f"Tasks in parallel_group '{group_name}' have a "
                    f"cross-dependency: {dep_from} → {dep_to}"
                )

    # 6. Orphan check — every task should be reachable from a root
    #    (a root is a task with no dependencies)
    children_of: Dict[str, List[str]] = {t.temp_id: [] for t in tasks}
    for dep in deps:
        children_of.setdefault(dep.from_task_temp_id, []).append(dep.to_task_temp_id)

    tasks_with_parents = {d.to_task_temp_id for d in deps}
    roots = task_ids - tasks_with_parents

    if not roots:
        errors.append("No root tasks found (all tasks have dependencies — possible cycle)")
    else:
        reachable: set = set()
        stack = list(roots)
        while stack:
            node = stack.pop()
            if node in reachable:
                continue
            reachable.add(node)
            stack.extend(children_of.get(node, []))

        orphans = task_ids - reachable
        if orphans:
            errors.append(
                f"Orphan tasks not reachable from any root: {sorted(orphans)}"
            )

    return errors
