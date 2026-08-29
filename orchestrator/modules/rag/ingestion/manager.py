
"""
Document Management System for Context Engineering
=================================================

This module provides comprehensive document management capabilities including:
- Multi-format document processing (PDF, DOCX, MD, TXT)
- Intelligent chunking strategies
- Embedding generation and storage
- Document metadata management
- Source attribution and tracking
"""

import asyncio
import hashlib
import json
import logging
import mimetypes
import os
import tempfile
import time
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple, Any
from dataclasses import dataclass, asdict
from enum import Enum

import magic
import pdfplumber
from docx import Document as DocxDocument
from langchain_text_splitters import (
    RecursiveCharacterTextSplitter,
    MarkdownTextSplitter,
    MarkdownHeaderTextSplitter,
    PythonCodeTextSplitter
)
from langchain_core.documents import Document
# OpenAI embeddings handled by centralized manager
import psycopg2
from botocore.exceptions import ClientError

from core.storage import FAST_FAIL, ensure_bucket, get_s3_client

# Use SemanticChunker from RAG module
try:
    from modules.rag.chunking.semantic_chunker import SemanticChunker, ChunkingStrategy
    SEMANTIC_CHUNKER_AVAILABLE = True
except ImportError:
    SEMANTIC_CHUNKER_AVAILABLE = False
from psycopg2.extras import RealDictCursor
import numpy as np

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class DocumentStatus(Enum):
    PENDING = "pending"
    PROCESSING = "processing"
    COMPLETED = "completed"
    FAILED = "failed"

class DocumentType(Enum):
    PDF = "pdf"
    DOCX = "docx"
    MARKDOWN = "md"
    TEXT = "txt"
    PYTHON = "py"
    JSON = "json"
    XLSX = "xlsx"
    CSV = "csv"

@dataclass
class DocumentMetadata:
    filename: str
    file_type: str
    file_size: int
    upload_date: datetime
    processed_date: Optional[datetime] = None
    status: DocumentStatus = DocumentStatus.PENDING
    chunk_count: int = 0
    tags: List[str] = None
    description: str = ""
    created_by: str = "system"
    
    def to_dict(self) -> Dict:
        data = asdict(self)
        data['upload_date'] = self.upload_date.isoformat()
        if self.processed_date:
            data['processed_date'] = self.processed_date.isoformat()
        data['status'] = self.status.value
        return data

@dataclass
class DocumentChunk:
    document_id: int
    chunk_index: int
    content: str
    embedding: Optional[List[float]] = None
    metadata: Dict = None
    parent_content: Optional[str] = None  # NEW: Parent section for context expansion
    headers: Dict[str, str] = None  # NEW: Header hierarchy (h1, h2, h3)
    
    def to_dict(self) -> Dict:
        return {
            'document_id': self.document_id,
            'chunk_index': self.chunk_index,
            'content': self.content,
            'metadata': self.metadata or {},
            'parent_content': self.parent_content,
            'headers': self.headers or {}
        }

class DocumentProcessor:
    """Handles processing of different document types"""
    
    def __init__(self):
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=1000,
            chunk_overlap=200,
            length_function=len,
        )
        self.markdown_splitter = MarkdownTextSplitter(
            chunk_size=1000,
            chunk_overlap=200
        )
        self.code_splitter = PythonCodeTextSplitter(
            chunk_size=1000,
            chunk_overlap=200
        )
    
    def detect_file_type(self, file_path: str) -> DocumentType:
        """Detect file type using magic numbers and extension"""
        try:
            mime_type = magic.from_file(file_path, mime=True)
            extension = Path(file_path).suffix.lower()
            
            if mime_type == 'application/pdf' or extension == '.pdf':
                return DocumentType.PDF
            elif mime_type in ['application/vnd.openxmlformats-officedocument.wordprocessingml.document'] or extension == '.docx':
                return DocumentType.DOCX
            elif extension in ['.md', '.markdown']:
                return DocumentType.MARKDOWN
            elif extension == '.py':
                return DocumentType.PYTHON
            elif extension == '.json':
                return DocumentType.JSON
            elif mime_type in ['application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'] or extension == '.xlsx':
                return DocumentType.XLSX
            elif mime_type in ['text/csv'] or extension == '.csv':
                return DocumentType.CSV
            else:
                return DocumentType.TEXT
        except Exception as e:
            logger.warning(f"Could not detect file type for {file_path}: {e}")
            return DocumentType.TEXT
    
    def extract_text_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF using pdfplumber with PyPDF2 fallback"""
        # Try pdfplumber first (best quality)
        try:
            text = ""
            with pdfplumber.open(file_path) as pdf:
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        # Remove null characters and other problematic characters
                        page_text = page_text.replace('\x00', '').replace('\x01', '').replace('\x02', '')
                        # Fix double characters (common PDF extraction issue)
                        import re
                        page_text = re.sub(r'(.)\1+', r'\1', page_text)  # Remove duplicate characters
                        text += page_text + "\n"
            if text.strip():
                return text.strip()
        except Exception as e:
            logger.warning(f"pdfplumber failed for {file_path}: {e}. Trying PyPDF2 fallback...")

        # Fallback to PyPDF2 (handles more PDF varieties)
        try:
            import PyPDF2
            text = ""
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                for page in pdf_reader.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
            if text.strip():
                logger.info(f"Successfully extracted text from PDF using PyPDF2 fallback: {file_path}")
                return text.strip()
        except Exception as e:
            logger.error(f"PyPDF2 also failed for {file_path}: {e}")

        # If both fail, raise error
        raise Exception(f"Could not extract text from PDF {file_path} with any available parser")
    
    def extract_text_from_docx(self, file_path: str) -> str:
        """Extract text from DOCX file"""
        try:
            doc = DocxDocument(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text.strip()
        except Exception as e:
            logger.error(f"Error extracting text from DOCX {file_path}: {e}")
            raise

    def _extract_spreadsheet_xlsx(self, file_path: str) -> str:
        """Extract XLSX spreadsheet data as Markdown tables for LLM consumption."""
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl not installed, cannot extract spreadsheet")
            return ""

        wb = openpyxl.load_workbook(file_path, data_only=True)
        parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            parts.append(f"## Sheet: {sheet_name}\n")

            headers = rows[0]
            md = "| " + " | ".join(str(h or "") for h in headers) + " |\n"
            md += "| " + " | ".join("---" for _ in headers) + " |\n"
            for row in rows[1:]:
                md += "| " + " | ".join(str(cell or "") for cell in row) + " |\n"

            parts.append(md)

        return "\n\n".join(parts)

    def _extract_spreadsheet_csv(self, file_path: str) -> str:
        """Extract CSV data as a Markdown table for LLM consumption."""
        import csv

        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                reader = csv.reader(f)
                rows = list(reader)
        except UnicodeDecodeError:
            with open(file_path, 'r', encoding='latin-1') as f:
                reader = csv.reader(f)
                rows = list(reader)

        if not rows:
            return ""

        headers = rows[0]
        md = "| " + " | ".join(str(h or "") for h in headers) + " |\n"
        md += "| " + " | ".join("---" for _ in headers) + " |\n"
        for row in rows[1:]:
            md += "| " + " | ".join(str(cell or "") for cell in row) + " |\n"

        return md

    def extract_text_from_file(self, file_path: str) -> str:
        """Extract text from any supported file type"""
        file_type = self.detect_file_type(file_path)

        if file_type == DocumentType.PDF:
            text = self.extract_text_from_pdf(file_path)
        elif file_type == DocumentType.DOCX:
            text = self.extract_text_from_docx(file_path)
        elif file_type == DocumentType.XLSX:
            text = self._extract_spreadsheet_xlsx(file_path)
        elif file_type == DocumentType.CSV:
            text = self._extract_spreadsheet_csv(file_path)
        else:
            # For text-based files (MD, TXT, PY, JSON)
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    text = f.read()
            except UnicodeDecodeError:
                # Try with different encoding
                with open(file_path, 'r', encoding='latin-1') as f:
                    text = f.read()
        
        # Clean any remaining null characters from all text sources
        if text:
            text = text.replace('\x00', '').replace('\x01', '').replace('\x02', '')
        
        return text
    
    def chunk_document(self, text: str, file_type: DocumentType, metadata: Dict = None) -> List[DocumentChunk]:
        """
        Split document into chunks using EXISTING SemanticChunker.
        
        Uses context_engineering/chunking/semantic_chunker.py which has:
        - SEMANTIC_SIMILARITY strategy
        - INFORMATION_DENSITY strategy  
        - TOPIC_COHERENCE strategy
        - HIERARCHICAL strategy
        - ADAPTIVE strategy
        
        Plus mathematical foundations (entropy, information theory).
        """
        chunks = []
        
        # USE EXISTING SEMANTIC CHUNKER (NOT A DUPLICATE!)
        if SEMANTIC_CHUNKER_AVAILABLE:
            try:
                # Use TOPIC_COHERENCE — fast keyword-based chunking (no local model)
                # ADAPTIVE runs 3 strategies with embedding calls = ~18s/batch on CPU
                semantic_chunker = SemanticChunker(
                    strategy=ChunkingStrategy.TOPIC_COHERENCE,
                    target_chunk_size=500,
                    min_chunk_size=100,
                    max_chunk_size=1500,
                    overlap_ratio=0.1,
                    similarity_threshold=0.7
                )
                semantic_chunker._use_embeddings = False  # Skip local model loading
                
                doc_id = metadata.get('document_id') if metadata else None
                semantic_chunks = semantic_chunker.chunk_text(text, document_id=str(doc_id) if doc_id else None)
                
                for i, sc in enumerate(semantic_chunks):
                    chunk_metadata = {
                        'file_type': file_type.value,
                        'chunk_size': len(sc.content),
                        # Preserve mathematical metrics from SemanticChunker
                        'entropy': sc.metadata.entropy,
                        'topic_coherence': sc.metadata.topic_coherence,
                        'semantic_density': sc.metadata.semantic_density,
                        'importance_score': sc.metadata.importance_score,
                        **(metadata or {})
                    }
                    
                    chunks.append(DocumentChunk(
                        document_id=0,
                        chunk_index=i,
                        content=sc.content,
                        metadata=chunk_metadata,
                        parent_content=None,  # SemanticChunker handles this differently
                        headers={}
                    ))
                
                logger.info(f"SemanticChunker (TOPIC_COHERENCE) created {len(chunks)} chunks with entropy/coherence metrics")
                return chunks
                
            except Exception as e:
                logger.warning(f"SemanticChunker failed, falling back to basic chunking: {e}")
                import traceback
                traceback.print_exc()
        
        # FALLBACK: Use basic splitters for non-markdown or if SmartChunker fails
        if file_type == DocumentType.PYTHON:
            docs = self.code_splitter.create_documents([text])
        else:
            docs = self.text_splitter.create_documents([text])
        
        # POST-PROCESS: Filter and merge chunks
        valid_chunks = []
        chunk_index = 0
        
        for doc in docs:
            content = doc.page_content.strip()
            
            # FILTER OUT BAD CHUNKS: separators, empty, too short
            if not content or len(content) < 50:
                continue
            
            # Skip chunks that are just separators or code fences
            if content in ['---', '```', '```python', '```javascript', '```typescript', '```json', '```yaml', '```bash']:
                continue
            
            # Skip chunks that are mostly separators/whitespace
            without_separators = content.replace('-', '').replace('=', '').replace('_', '').replace('#', '').replace('`', '').replace('*', '').strip()
            if len(without_separators) < 30:
                continue
            
            # Skip chunks that are just markdown headers with no content
            if content.count('\n') == 0 and content.startswith('#'):
                continue
            
            # Check if chunk has meaningful content (not just formatting)
            words = without_separators.split()
            if len(words) < 5:
                continue
            
            chunk_metadata = {
                'file_type': file_type.value,
                'chunk_size': len(content),
                **(metadata or {})
            }
            
            valid_chunks.append(DocumentChunk(
                document_id=0,
                chunk_index=chunk_index,
                content=content,
                metadata=chunk_metadata
            ))
            chunk_index += 1
        
        return valid_chunks

class DocumentManager:
    """Main document management class"""

    def __init__(
        self,
        db_config: Dict[str, str],
        workspace_id: Optional[str] = None,
        use_s3_vectors: bool = False,
        s3_bucket: Optional[str] = None
    ):
        self.db_config = db_config
        self.processor = DocumentProcessor()
        self._db_initialized = False
        self._init_lock = False  # Simple flag to prevent concurrent initialization
        self.workspace_id = workspace_id
        self.use_s3_vectors = use_s3_vectors
        self._s3_backend = None

        # S3 configuration for document storage. The client comes from the
        # core.storage factory (PRD-233 S4; honours S3_ENDPOINT_URL for MinIO)
        # on FIRST USE — constructing a DocumentManager performs no network I/O.
        # The FAST_FAIL profile keeps the PRD-164 contract: bounded timeouts and
        # a single attempt, so a slow/unreachable/unconfigured object store
        # fails a document op in seconds instead of hanging the process (this
        # was a 90s faulthandler hang in CI when no creds were present).
        from config import config as app_config
        self.s3_bucket = s3_bucket or app_config.S3_DOCUMENTS_BUCKET

        logger.info(f"✅ DocumentManager initialized with S3 document storage: bucket={self.s3_bucket}")

        # Initialize S3 vectors backend if enabled
        if self.use_s3_vectors:
            if not self.workspace_id:
                raise ValueError("workspace_id required when use_s3_vectors=True")
            from modules.search.vector_store import get_vector_store
            self._s3_backend = get_vector_store(
                backend="s3_vectors",
                workspace_id=self.workspace_id
            )
            logger.info(f"✅ DocumentManager using S3 vectors for workspace {workspace_id}")
        
        # Use centralized embedding manager
        from core.llm import create_embedding_manager
        self.embedding_manager = create_embedding_manager()
        logger.info(f"DocumentManager using {self.embedding_manager.get_provider_info()['provider']} embeddings")
        
        # Don't initialize database at import time - do it lazily on first use

    @property
    def s3_client(self):
        """The process-wide fast-fail S3 client (core.storage) — built on first use."""
        return get_s3_client(FAST_FAIL)

    def _ensure_s3_bucket_exists(self):
        """Self-create the documents bucket before the first write.

        Runs lazily (first upload), not at construction. Only creates against a
        configured endpoint (MinIO); on AWS the bucket is infrastructure and
        ``ensure_bucket`` is a no-op. A failed check must NOT crash the write
        path here — the put that follows surfaces the precise error itself.
        """
        try:
            ensure_bucket(self.s3_bucket, self.s3_client)
        except Exception as e:
            logger.warning(f"⚠️ S3 bucket check skipped (S3 unavailable — continuing): {e}")

    def _extract_pdf_with_tables(self, file_path: str) -> tuple:
        """
        Extract text AND tables from PDF.
        Tables are converted to Markdown format for better LLM comprehension.
        Returns: (full_text, tables_list)
        """
        import pdfplumber

        text_parts = []
        tables = []

        with pdfplumber.open(file_path) as pdf:
            for page_num, page in enumerate(pdf.pages, 1):
                # Extract regular text
                page_text = page.extract_text() or ""
                text_parts.append(f"[Page {page_num}]\n{page_text}")

                # Extract tables
                page_tables = page.extract_tables()
                for table_idx, table in enumerate(page_tables):
                    if table and len(table) > 1:
                        headers = table[0]
                        md_table = "| " + " | ".join(str(h or "") for h in headers) + " |\n"
                        md_table += "| " + " | ".join("---" for _ in headers) + " |\n"
                        for row in table[1:]:
                            md_table += "| " + " | ".join(str(cell or "") for cell in row) + " |\n"

                        text_parts.append(f"\n[Table {table_idx + 1}, Page {page_num}]\n{md_table}")
                        tables.append({
                            "page": page_num,
                            "index": table_idx,
                            "markdown": md_table,
                        })

        return "\n\n".join(text_parts), tables

    def _extract_spreadsheet(self, file_path: str) -> str:
        """Extract spreadsheet data as Markdown tables for LLM consumption."""
        try:
            import openpyxl
        except ImportError:
            logger.warning("openpyxl not installed, cannot extract spreadsheet")
            return ""

        wb = openpyxl.load_workbook(file_path, data_only=True)
        parts = []

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            parts.append(f"## Sheet: {sheet_name}\n")

            headers = rows[0]
            md = "| " + " | ".join(str(h or "") for h in headers) + " |\n"
            md += "| " + " | ".join("---" for _ in headers) + " |\n"
            for row in rows[1:]:
                md += "| " + " | ".join(str(cell or "") for cell in row) + " |\n"

            parts.append(md)

        return "\n\n".join(parts)

    def _ensure_database_initialized(self, max_retries: int = 5, retry_delay: float = 2.0):
        """Ensure database is initialized with retry logic"""
        if self._db_initialized:
            return
        
        if self._init_lock:
            # Wait for concurrent initialization
            for _ in range(max_retries):
                time.sleep(retry_delay)
                if self._db_initialized:
                    return
            raise RuntimeError("Database initialization timeout")
        
        self._init_lock = True
        try:
            for attempt in range(max_retries):
                try:
                    self._init_database()
                    self._db_initialized = True
                    logger.info("Database initialized successfully")
                    return
                except (psycopg2.OperationalError, psycopg2.InterfaceError) as e:
                    if "the database system is starting up" in str(e) or "connection" in str(e).lower():
                        if attempt < max_retries - 1:
                            wait_time = retry_delay * (2 ** attempt)  # Exponential backoff
                            logger.warning(f"Database not ready (attempt {attempt + 1}/{max_retries}), retrying in {wait_time}s...")
                            time.sleep(wait_time)
                            continue
                    raise
            raise RuntimeError(f"Failed to initialize database after {max_retries} attempts")
        finally:
            self._init_lock = False
    
    def _init_database(self):
        """Initialize database tables if they don't exist"""
        conn = psycopg2.connect(**self.db_config)
        try:
            cursor = conn.cursor()
            
            # Create documents table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS documents (
                    id SERIAL PRIMARY KEY,
                    filename VARCHAR(255) NOT NULL,
                    file_type VARCHAR(50) NOT NULL,
                    file_size INTEGER NOT NULL,
                    upload_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    processed_date TIMESTAMP,
                    status VARCHAR(50) DEFAULT 'pending',
                    chunk_count INTEGER DEFAULT 0,
                    metadata JSONB DEFAULT '{}',
                    created_by VARCHAR(100) DEFAULT 'system',
                    tags TEXT[] DEFAULT ARRAY[]::TEXT[],
                    description TEXT DEFAULT '',
                    file_hash VARCHAR(64) UNIQUE,
                    workspace_id TEXT,
                    source_type VARCHAR(50)
                );
            """)
            
            # Create document_chunks table
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS document_chunks (
                    id SERIAL PRIMARY KEY,
                    document_id INTEGER REFERENCES documents(id) ON DELETE CASCADE,
                    chunk_index INTEGER NOT NULL,
                    content TEXT NOT NULL,
                    embedding TEXT,
                    metadata JSONB DEFAULT '{}',
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    workspace_id TEXT,
                    parent_content TEXT,
                    headers JSONB DEFAULT '{}'
                );
            """)
            
            # Create context_usage table for analytics
            cursor.execute("""
                CREATE TABLE IF NOT EXISTS context_usage (
                    id SERIAL PRIMARY KEY,
                    document_id INTEGER REFERENCES documents(id),
                    chunk_id INTEGER REFERENCES document_chunks(id),
                    query_text TEXT,
                    relevance_score FLOAT,
                    used_in_response BOOLEAN DEFAULT FALSE,
                    timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP
                );
            """)
            
            # Create indexes for better performance
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_document_chunks_document_id ON document_chunks(document_id);")
            # cursor.execute("CREATE INDEX IF NOT EXISTS idx_document_chunks_embedding ON document_chunks USING ivfflat (embedding vector_cosine_ops);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_documents_status ON documents(status);")
            cursor.execute("CREATE INDEX IF NOT EXISTS idx_documents_file_type ON documents(file_type);")
            
            conn.commit()
        finally:
            cursor.close()
            conn.close()
    
    def _calculate_file_hash(self, file_path: str) -> str:
        """Calculate SHA-256 hash of file for deduplication"""
        hash_sha256 = hashlib.sha256()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_sha256.update(chunk)
        return hash_sha256.hexdigest()

    def _upload_to_s3(self, file_path: str, document_id: int, filename: str) -> str:
        """
        Upload document to S3 with workspace isolation.

        Returns S3 key (path) for the uploaded file.
        Structure: workspaces/{workspace_id}/documents/{document_id}_{filename}
        """
        if not self.workspace_id:
            raise ValueError("workspace_id required for S3 upload")

        # S3 key with workspace isolation
        s3_key = f"workspaces/{self.workspace_id}/documents/{document_id}_{filename}"

        self._ensure_s3_bucket_exists()
        try:
            with open(file_path, 'rb') as f:
                self.s3_client.put_object(
                    Bucket=self.s3_bucket,
                    Key=s3_key,
                    Body=f,
                    ContentType=mimetypes.guess_type(filename)[0] or 'application/octet-stream'
                )
            logger.info(f"✅ Uploaded document to S3: s3://{self.s3_bucket}/{s3_key}")
            return s3_key
        except ClientError as e:
            logger.error(f"Failed to upload document to S3: {e}")
            raise

    def _download_from_s3(self, s3_key: str, local_path: str) -> None:
        """Download document from S3 to local path for processing."""
        try:
            self.s3_client.download_file(self.s3_bucket, s3_key, local_path)
            logger.info(f"✅ Downloaded from S3: {s3_key} → {local_path}")
        except ClientError as e:
            logger.error(f"Failed to download from S3: {e}")
            raise
    
    async def upload_document(self, file_path: str, filename: str = None,
                            tags: List[str] = None, description: str = "",
                            created_by: str = "system",
                            source_type: Optional[str] = None) -> int:
        """Upload and process a document.

        ``source_type`` (PRD-164 S3, Q58): provenance scope. ``None`` for a
        regular upload; ``'agent_output'`` when the knowledge flywheel routes
        an agent output (mission synthesis / generated document / report)
        through this manager.
        """
        self._ensure_database_initialized()
        try:
            if not os.path.exists(file_path):
                raise FileNotFoundError(f"File not found: {file_path}")
            
            if filename is None:
                filename = os.path.basename(file_path)

            # Calculate file metadata from local temp file
            file_size = os.path.getsize(file_path)
            file_hash = self._calculate_file_hash(file_path)
            file_type = self.processor.detect_file_type(file_path)
            
            # Check if document already exists
            conn = psycopg2.connect(**self.db_config)
            cursor = conn.cursor()
            
            cursor.execute("SELECT id, status FROM documents WHERE file_hash = %s", (file_hash,))
            existing = cursor.fetchone()
            if existing:
                existing_id, existing_status = existing
                # Only return existing document if it was successfully processed
                if existing_status == DocumentStatus.COMPLETED.value:
                    logger.info(f"Document with hash {file_hash} already exists with ID {existing_id} (status: {existing_status})")
                    cursor.close()
                    conn.close()
                    return existing_id
                else:
                    # Document exists but failed/pending - delete and re-process
                    logger.warning(f"Document with hash {file_hash} exists with ID {existing_id} but has status '{existing_status}'. Deleting and re-processing...")
                    cursor.execute("DELETE FROM document_chunks WHERE document_id = %s", (existing_id,))
                    cursor.execute("DELETE FROM documents WHERE id = %s", (existing_id,))
                    conn.commit()
                    cursor.close()
                    conn.close()
                    logger.info(f"Deleted failed document {existing_id}, will re-process")

                    # Reopen connection for new document creation
                    conn = psycopg2.connect(**self.db_config)
                    cursor = conn.cursor()

            # Create document record (without file_path yet)
            metadata = DocumentMetadata(
                filename=filename,
                file_type=file_type.value,
                file_size=file_size,
                upload_date=datetime.now(),
                tags=tags or [],
                description=description,
                created_by=created_by
            )

            cursor.execute("""
                INSERT INTO documents (filename, file_type, file_size, upload_date, status,
                                     metadata, created_by, tags, description, file_hash, workspace_id,
                                     source_type)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            """, (
                metadata.filename, metadata.file_type, metadata.file_size,
                metadata.upload_date, DocumentStatus.PROCESSING.value,
                json.dumps(metadata.to_dict()), metadata.created_by,
                metadata.tags, metadata.description, file_hash, self.workspace_id,
                source_type
            ))

            document_id = cursor.fetchone()[0]
            conn.commit()

            # Upload to S3 with workspace isolation
            s3_key = self._upload_to_s3(file_path, document_id, filename)

            # Update document with S3 path
            cursor.execute("""
                UPDATE documents SET file_path = %s WHERE id = %s
            """, (f"s3://{self.s3_bucket}/{s3_key}", document_id))
            conn.commit()

            # Process document (it will read from local temp file)
            await self._process_document(document_id, file_path, file_type, s3_key, filename=filename)
            
            cursor.close()
            conn.close()
            
            logger.info(f"Document uploaded successfully with ID: {document_id}")
            return document_id
            
        except Exception as e:
            logger.error(f"Error uploading document: {e}")
            raise
    
    async def _process_document(self, document_id: int, file_path: str, file_type: DocumentType, s3_key: Optional[str] = None, filename: str = None):
        """
        Process document: extract text, chunk, and generate embeddings.

        Args:
            document_id: Database document ID
            file_path: Local temp file path for processing
            file_type: Document type
            s3_key: S3 key where document is stored (for reference)
            filename: Real document filename (not the temp path basename)
        """
        self._ensure_database_initialized()
        # W3-S8: track S3 vector_ids the persist helper stored so the outer
        # except can clean up if a later step (status update, commit) fails
        # after the dual-write already succeeded.
        vector_ids: List[str] = []
        conn = None
        try:
            logger.info(f"Starting document processing for document {document_id}, type: {file_type} (S3: {s3_key})")

            conn = psycopg2.connect(**self.db_config)
            cursor = conn.cursor()

            # Fetch workspace_id and team_access from document record
            cursor.execute("SELECT workspace_id, team_access FROM documents WHERE id = %s", (document_id,))
            doc_row = cursor.fetchone()
            workspace_id = doc_row[0] if doc_row else None
            doc_team_access = doc_row[1] if doc_row and doc_row[1] else []
            logger.info(f"Processing document {document_id} for workspace {workspace_id}")

            # Extract text (with enhanced extraction for PDFs and spreadsheets)
            pdf_tables = []
            if file_type == DocumentType.PDF:
                try:
                    text, pdf_tables = self._extract_pdf_with_tables(file_path)
                    logger.info(f"Extracted {len(text)} characters and {len(pdf_tables)} tables from PDF document {document_id}")
                except Exception as e:
                    logger.warning(f"Enhanced PDF extraction failed, falling back to basic: {e}")
                    text = self.processor.extract_text_from_file(file_path)
            elif file_type == DocumentType.XLSX:
                try:
                    text = self._extract_spreadsheet(file_path)
                    logger.info(f"Extracted {len(text)} characters from XLSX document {document_id}")
                except Exception as e:
                    logger.warning(f"XLSX extraction failed: {e}")
                    text = self.processor.extract_text_from_file(file_path)
            else:
                text = self.processor.extract_text_from_file(file_path)
            logger.info(f"Extracted {len(text)} characters from document {document_id}")
            
            # Multimodal processing (tables, formulas, images)
            try:
                from modules.rag.ingestion.multimodal import TableProcessor, FormulaProcessor, ImageProcessor
                from modules.search.services.entity_extractor import EntityExtractor, create_or_get_entity, create_entity_mention, create_entity_relationship
                
                logger.info(f"Starting multimodal processing for document {document_id}")
                
                # Create knowledge_items entry for this document (if not exists)
                cursor.execute("""
                    INSERT INTO knowledge_items (id, kb_type_id, title, content, metadata, quality_score, workspace_id)
                    SELECT %s,
                           (SELECT id FROM kb_types WHERE type_name = 'document' LIMIT 1),
                           %s,
                           %s,
                           %s,
                           0.8,
                           %s
                    WHERE NOT EXISTS (SELECT 1 FROM knowledge_items WHERE id = %s)
                """, (
                    document_id,
                    os.path.basename(file_path),
                    text[:1000] if text else '',  # First 1000 chars as preview
                    json.dumps({'source': 'document_upload', 'file_type': str(file_type)}),
                    workspace_id,
                    document_id
                ))
                conn.commit()
                logger.info(f"Created knowledge_items entry for document {document_id}")
                
                # Extract tables based on file type
                tables = []
                try:
                    if file_type == DocumentType.PDF or file_type == DocumentType.DOCX:
                        # Use camelot for PDFs
                        table_processor = TableProcessor()
                        tables = table_processor.extract_tables_from_pdf(file_path)
                    elif file_type == DocumentType.MARKDOWN or file_type == DocumentType.TEXT:
                        # Extract Markdown tables using regex
                        import re
                        
                        # Find Markdown table patterns
                        table_pattern = r'\|(.+)\|[\r\n]+\|[\s:-]+\|[\r\n]+((?:\|.+\|[\r\n]+)+)'
                        matches = re.finditer(table_pattern, text, re.MULTILINE)
                        
                        for match in matches:
                            table_text = match.group(0)
                            # Convert to simple format
                            tables.append({
                                'markdown': table_text,
                                'data': [],  # Could parse if needed
                                'csv': table_text.replace('|', ',')  # Simple CSV conversion
                            })
                        
                        logger.info(f"Found {len(tables)} markdown tables in document {document_id}")
                    
                    # Insert all extracted tables (from PDF or Markdown)
                    if tables:
                        for idx, table_data in enumerate(tables):
                            # Handle both dict and dataclass formats
                            if isinstance(table_data, dict):
                                markdown_content = table_data.get('markdown')
                                json_content = json.dumps(table_data.get('data', []))
                                csv_content = table_data.get('csv')
                            else:
                                # TableExtraction dataclass
                                markdown_content = table_data.markdown
                                json_content = json.dumps(table_data.json)
                                csv_content = table_data.csv
                            
                            # Extract table details
                            if hasattr(table_data, 'headers'):
                                headers = json.dumps(table_data.headers)
                                row_count = table_data.row_count
                                column_count = table_data.column_count
                            else:
                                headers = json.dumps([])
                                row_count = len(markdown_content.split('\n')) if markdown_content else 0
                                column_count = len(markdown_content.split('|')) if markdown_content else 0
                            
                            cursor.execute("""
                                INSERT INTO kb_tables (
                                    knowledge_item_id, headers, row_count, column_count, 
                                    markdown_representation, csv_data, json_data
                                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                            """, (
                                document_id,
                                headers,
                                row_count,
                                column_count,
                                markdown_content,
                                csv_content,
                                json_content
                            ))
                        
                        conn.commit()
                        logger.info(f"Inserted {len(tables)} tables into knowledge base for document {document_id}")
                    
                except Exception as e:
                    logger.warning(f"Table extraction failed for document {document_id}: {e}")
                
                # Process formulas (for ALL file types)
                try:
                    formula_processor = FormulaProcessor()
                    formulas = formula_processor.extract_formulas_from_text(text)
                    
                    if formulas:
                        for idx, formula_data in enumerate(formulas):
                            # Handle both dict and dataclass formats
                            if isinstance(formula_data, dict):
                                latex_content = formula_data.get('latex')
                                mathml_content = formula_data.get('mathml')
                                rendered_url = formula_data.get('rendered_url')
                            else:
                                # FormulaExtraction dataclass
                                latex_content = formula_data.latex
                                mathml_content = None  # Not provided by FormulaExtraction
                                rendered_url = None    # Not provided by FormulaExtraction
                            
                            cursor.execute("""
                                INSERT INTO kb_formulas (
                                    knowledge_item_id, latex, mathml, ascii_math, workspace_id
                                ) VALUES (%s, %s, %s, %s, %s)
                            """, (
                                document_id,
                                latex_content,
                                mathml_content or '',
                                formula_data.ascii_math if hasattr(formula_data, 'ascii_math') else '',
                                workspace_id
                            ))
                        
                        conn.commit()
                        logger.info(f"Inserted {len(formulas)} formulas into knowledge base for document {document_id}")
                
                except Exception as e:
                    logger.warning(f"Formula extraction failed for document {document_id}: {e}")
                
                # NEW: Entity extraction for knowledge graph
                try:
                    logger.info(f"Starting entity extraction for document {document_id}")
                    entity_extractor = EntityExtractor()
                    
                    # Extract entities from text
                    entities = await entity_extractor.extract_entities(text, use_llm=True, max_entities=30)
                    logger.info(f"Extracted {len(entities)} entities from document {document_id}")
                    
                    # Store entities and create mentions
                    entity_ids = []
                    for entity in entities:
                        entity_id = await create_or_get_entity(
                            cursor,
                            entity.entity_name,
                            entity.entity_type,
                            entity.canonical_name,
                            entity.description,
                            workspace_id=workspace_id
                        )
                        entity_ids.append(entity_id)
                        
                        await create_entity_mention(
                            cursor,
                            document_id,
                            entity_id,
                            entity.mention_context,
                            entity.confidence,
                            entity.position,
                            "llm"
                        )
                    
                    conn.commit()
                    logger.info(f"Stored {len(entities)} entities for document {document_id}")
                    
                    # Extract relationships between entities
                    if len(entities) >= 2:
                        relationships = await entity_extractor.extract_relationships(text, entities, max_relationships=20)
                        logger.info(f"Extracted {len(relationships)} relationships from document {document_id}")
                        
                        for rel in relationships:
                            await create_entity_relationship(
                                cursor,
                                rel.from_entity,
                                rel.to_entity,
                                rel.relationship_type,
                                rel.strength,
                                document_id,
                                rel.evidence,
                                workspace_id=workspace_id
                            )
                        
                        conn.commit()
                        logger.info(f"Stored {len(relationships)} relationships for document {document_id}")
                    
                except Exception as e:
                    logger.warning(f"Entity extraction failed for document {document_id}: {e}")
                    import traceback
                    traceback.print_exc()
                
                conn.commit()
                
            except Exception as e:
                logger.warning(f"Multimodal processing encountered errors for document {document_id}: {e}")
                # Continue processing even if multimodal extraction fails
            
            # Optional multimodal processing for PDFs: extract tables/images as additional chunks
            additional_chunks = []
            if file_type == DocumentType.PDF and getattr(self, 'enable_multimodal', False):
                try:
                    from modules.rag.ingestion.multimodal.processors import create_multimodal_processor
                    mm_processor = create_multimodal_processor(openai_key=getattr(self, 'openai_key', None))
                    multimodal_results = mm_processor.process_pdf_multimodal(file_path)

                    for table in multimodal_results.get('tables', []):
                        markdown = table.markdown if hasattr(table, 'markdown') else table.get('markdown', '')
                        page = table.page_number if hasattr(table, 'page_number') else table.get('page')
                        if markdown:
                            additional_chunks.append({
                                "content": markdown,
                                "metadata": {"type": "table", "page": page, "source": "multimodal"}
                            })

                    for image in multimodal_results.get('images', []):
                        description = image.description if hasattr(image, 'description') else image.get('description')
                        page = image.page_number if hasattr(image, 'page_number') else image.get('page')
                        if description:
                            additional_chunks.append({
                                "content": f"[Image: {description}]",
                                "metadata": {"type": "image", "page": page, "source": "multimodal"}
                            })
                except Exception as e:
                    logger.warning(f"Multimodal processing failed (non-fatal): {e}")

            # Create chunks with file_path metadata
            chunks = self.processor.chunk_document(text, file_type, {
                'document_id': document_id,
                'source_file': os.path.basename(file_path),
                'file_path': file_path,  # NEW: Add full path for downloads
                'team_access': doc_team_access,  # PRD-124: team scoping in chunk metadata
            })

            # Append multimodal chunks to the regular chunks
            for mc in additional_chunks:
                chunks.append(DocumentChunk(
                    document_id=document_id,
                    chunk_index=len(chunks),
                    content=mc["content"],
                    metadata={
                        'document_id': document_id,
                        'source_file': os.path.basename(file_path),
                        'file_path': file_path,
                        **mc["metadata"]
                    }
                ))

            # Phase 1: Filter and enrich chunks (no I/O)
            filtered_chunks = []
            for chunk in chunks:
                chunk.document_id = document_id

                # SAFETY CHECK: Skip bad chunks (separators, empty, too short)
                content = chunk.content.strip() if chunk.content else ""
                if not content or len(content) < 20:
                    continue
                if content in ['---', '```', '```python', '```javascript', '```typescript']:
                    continue
                stripped = content.replace('-', '').replace('=', '').replace('_', '').replace('#', '').replace('`', '').strip()
                if len(stripped) < 10:
                    continue

                # Enrich chunk metadata with page numbers from [Page N] markers (PDF)
                if file_type == DocumentType.PDF:
                    import re
                    page_markers = re.findall(r'\[Page (\d+)\]', chunk.content)
                    if page_markers:
                        page_numbers = sorted(set(int(p) for p in page_markers))
                        if chunk.metadata is None:
                            chunk.metadata = {}
                        chunk.metadata['page_numbers'] = page_numbers
                        chunk.metadata['page_start'] = page_numbers[0]
                        chunk.metadata['page_end'] = page_numbers[-1]

                    # Check for table markers too
                    table_markers = re.findall(r'\[Table (\d+), Page (\d+)\]', chunk.content)
                    if table_markers:
                        if chunk.metadata is None:
                            chunk.metadata = {}
                        chunk.metadata['contains_tables'] = True
                        chunk.metadata['table_refs'] = [
                            {"table": int(t), "page": int(p)} for t, p in table_markers
                        ]

                filtered_chunks.append(chunk)

            # PRD-188 S2: contextual annotations — situate each chunk within
            # its parent document BEFORE embedding (Anthropic contextual-
            # retrieval pattern). The annotated content is what gets embedded
            # AND stored, so the tsvector trigger indexes the preface for the
            # BM25 leg too. Flag default OFF until the corpus reprocess has
            # run; failure is loud and falls back to raw text — an annotation
            # can never fail an ingest.
            if filtered_chunks:
                from config import config as app_config
                if app_config.RAG_CONTEXTUAL_ANNOTATIONS_ENABLED:
                    from modules.rag.ingestion.contextual_annotator import annotate_chunks
                    filtered_chunks = await annotate_chunks(filtered_chunks, text)

            # Phase 2: Batch embed all chunks with parallel processing
            if filtered_chunks:
                chunk_texts = [c.content for c in filtered_chunks]
                logger.info(f"Generating embeddings for {len(chunk_texts)} chunks (parallel batch)")

                try:
                    # Use batch embedding (parallel API calls)
                    all_embeddings = await self._generate_embeddings_batch(chunk_texts)
                except Exception as e:
                    logger.warning(f"Batch embedding failed ({e}), falling back to sequential")
                    all_embeddings = []
                    for text in chunk_texts:
                        emb = await self._generate_embedding(text)
                        all_embeddings.append(emb)

                for chunk, embedding in zip(filtered_chunks, all_embeddings):
                    chunk.embedding = embedding

            # Phase 3+4: dual-write chunks (postgres) + vectors (S3, if enabled).
            # §H "atomic per document": if S3 add_documents raises, the helper
            # rolls back the in-flight chunk INSERTs and re-raises — no
            # half-indexed doc survives the failure path.
            vector_ids = await self._persist_chunks_and_vectors(
                conn=conn,
                cursor=cursor,
                document_id=document_id,
                workspace_id=workspace_id,
                filename=filename,
                file_path=file_path,
                file_type=file_type,
                filtered_chunks=filtered_chunks,
            )
            valid_chunks = list(filtered_chunks)

            # Update document status
            cursor.execute("""
                UPDATE documents
                SET status = %s, processed_date = %s, chunk_count = %s
                WHERE id = %s
            """, (
                DocumentStatus.COMPLETED.value, datetime.now(), len(valid_chunks), document_id
            ))

            conn.commit()
            cursor.close()
            conn.close()

            logger.info(f"Document {document_id} processed successfully with {len(chunks)} chunks")

            # W3-S1 wiring: RAG primitive emits 'green' on a successful turn —
            # this is the only honest source of the RAG tile.
            self._emit_ingest_heartbeat(
                document_id=document_id,
                status="green",
                detail=f"doc {document_id} ingest ok ({len(valid_chunks)} chunks)",
            )

            # PRD-126: Trigger knowledge graph update on document ingest
            try:
                from modules.knowledge.graph_service import get_graph_service
                if self.workspace_id:
                    get_graph_service().schedule_incremental_update(
                        str(self.workspace_id),
                        [{"type": "document", "path": file_path, "id": document_id}],
                    )
            except Exception as graph_err:
                logger.warning("Graph update skipped: %s", graph_err)

        except Exception as e:
            logger.error(f"Error processing document {document_id}: {e}")

            # Explicit rollback of the in-flight conn — don't rely on GC for
            # atomicity (CPython detail, not a contract). The persist helper
            # already rolled back on its own S3 failure; this catches the
            # earlier-multimodal / status-update paths.
            try:
                if 'conn' in locals() and conn is not None:
                    conn.rollback()
            except Exception:
                logger.error(
                    f"Rollback failed for doc {document_id}", exc_info=True
                )

            # If the dual-write got past S3 but a *later* step (status update,
            # commit) failed, clean up the orphan vectors so the failed doc
            # doesn't leave retrievable chunks in the vector index.
            if vector_ids and self.use_s3_vectors and self._s3_backend:
                try:
                    self._s3_backend.delete_documents(str(document_id))
                    logger.info(
                        f"Cleaned up {len(vector_ids)} orphan S3 vectors after "
                        f"doc {document_id} ingest failure"
                    )
                except Exception:
                    logger.error(
                        f"S3 vector cleanup failed for doc {document_id} "
                        f"after ingest failure", exc_info=True
                    )

            # Update status to failed
            try:
                conn = psycopg2.connect(**self.db_config)
                cursor = conn.cursor()
                cursor.execute("""
                    UPDATE documents SET status = %s WHERE id = %s
                """, (DocumentStatus.FAILED.value, document_id))
                conn.commit()
                cursor.close()
                conn.close()
            except Exception:
                pass

            # W3-S1 wiring: RAG primitive emits 'down' on a failed ingest.
            self._emit_ingest_heartbeat(
                document_id=document_id,
                status="down",
                detail=f"doc {document_id} ingest failed: {type(e).__name__}",
            )

            raise

    async def _persist_chunks_and_vectors(
        self,
        conn,
        cursor,
        document_id: int,
        workspace_id: Optional[str],
        filename: Optional[str],
        file_path: str,
        file_type,
        filtered_chunks: List["DocumentChunk"],
    ) -> List[str]:
        """Persist chunks to Postgres and (in S3 Vectors mode) embeddings to S3.

        W3-S8 §H "atomic per document": the chunk INSERTs and the S3 batch
        ``add_documents`` call are a single logical unit. If the S3 step
        raises, the helper EXPLICITLY ``conn.rollback()``s before re-raising
        so the chunk INSERTs don't survive without their vector partners.
        Previously the code relied on garbage-collection rollback, which is a
        CPython implementation detail, not an atomicity contract.

        Returns the list of S3 vector_ids the backend stored, so the caller
        can clean them up if a *later* step (status update, commit, graph
        hook) fails after this helper returned successfully. Returns ``[]``
        in legacy pgvector mode (vectors live next to the chunks in
        Postgres, so the commit/rollback is symmetric automatically).
        """
        embeddings_for_s3: List[List[float]] = []
        documents_for_s3: List[Dict[str, Any]] = []

        for chunk in filtered_chunks:
            embedding = chunk.embedding
            chunk_headers = chunk.headers if getattr(chunk, "headers", None) else None

            if self.use_s3_vectors and self._s3_backend:
                # Store metadata in postgres WITHOUT embedding (S3 vectors mode)
                cursor.execute(
                    """
                    INSERT INTO document_chunks (document_id, chunk_index, content, metadata, parent_content, headers, workspace_id)
                    VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        chunk.document_id,
                        chunk.chunk_index,
                        chunk.content,
                        json.dumps(chunk.metadata),
                        chunk.parent_content,
                        json.dumps(chunk_headers) if chunk_headers else '{}',
                        workspace_id,
                    ),
                )

                embeddings_for_s3.append(embedding)
                real_name = filename or os.path.basename(file_path)
                documents_for_s3.append({
                    "external_file_id": str(document_id),
                    "document_id": str(document_id),
                    "chunk_index": chunk.chunk_index,
                    "chunk_text": chunk.content[:500],
                    "file_name": real_name,
                    "source_file": real_name,
                    "file_path": file_path,
                    "file_type": file_type.value if hasattr(file_type, 'value') else str(file_type),
                    "app_name": "document_sync",
                    "workspace_id": workspace_id,
                })
            else:
                # Legacy pgvector mode: chunk INSERT carries the embedding inline.
                cursor.execute(
                    """
                    INSERT INTO document_chunks (document_id, chunk_index, content, embedding, metadata, parent_content, headers, workspace_id)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                    """,
                    (
                        chunk.document_id,
                        chunk.chunk_index,
                        chunk.content,
                        embedding,
                        json.dumps(chunk.metadata),
                        chunk.parent_content,
                        json.dumps(chunk_headers) if chunk_headers else '{}',
                        workspace_id,
                    ),
                )

        if not (self.use_s3_vectors and self._s3_backend and embeddings_for_s3):
            return []

        try:
            await self._s3_backend.initialize()
            vector_ids = self._s3_backend.add_documents(
                documents=documents_for_s3,
                embeddings=embeddings_for_s3,
            )
            logger.info(
                f"✅ Stored {len(vector_ids)} vectors in S3 for document {document_id}"
            )
            return list(vector_ids or [])
        except Exception as e:
            logger.error(
                f"Failed to store vectors in S3 for doc {document_id}; "
                f"rolling back chunks: {e}"
            )
            try:
                conn.rollback()
            except Exception:
                logger.error(
                    f"Rollback after S3 failure for doc {document_id} also failed",
                    exc_info=True,
                )
            raise

    def _emit_ingest_heartbeat(
        self, document_id: int, status: str, detail: str = ""
    ) -> None:
        """W3-S1 wiring: best-effort RAG primitive_check emit.

        Never raises — a heartbeat write failure must not surface as an
        ingest failure (the §H 'observable' line is best-effort by design;
        mirrors the same swallow guarantee in ``emit_primitive_finding``).
        Silently skips when no ``workspace_id`` is in play — no per-workspace
        tile to update, and A4 forbids fabricating a default.
        """
        try:
            if not self.workspace_id:
                return
            from services.heartbeat_service import emit_primitive_finding

            emit_primitive_finding(
                str(self.workspace_id), "rag", status, detail
            )
        except Exception:
            logger.error(
                f"RAG heartbeat emit failed for doc {document_id}",
                exc_info=True,
            )
    
    def _get_target_dimension(self) -> int:
        """Get the configured embedding dimension for validation."""
        try:
            return self.embedding_manager.get_dimension()
        except Exception:
            from config import config as app_config
            return int(app_config.S3_VECTORS_DIMENSION or 2048)

    def _ensure_embedding_dimension(self, embedding: List[float]) -> List[float]:
        """Truncate embedding to target dimension if needed (Matryoshka support)."""
        target = self._get_target_dimension()
        if len(embedding) > target:
            return embedding[:target]
        return embedding

    async def _generate_embedding(self, text: str) -> List[float]:
        """Generate embedding for text with caching"""
        try:
            # Try cache first (massive cost savings!)
            from core.cache import get_cache_service
            cache = get_cache_service()
            target_dim = self._get_target_dimension()

            cached_embedding = cache.get_embedding(text)
            if cached_embedding is not None:
                # Validate dimension — stale cache may have wrong size
                if len(cached_embedding) == target_dim:
                    logger.debug(f"✅ Using cached embedding (saved API call)")
                    return cached_embedding
                elif len(cached_embedding) > target_dim:
                    # Matryoshka: truncate and re-cache
                    truncated = cached_embedding[:target_dim]
                    cache.set_embedding(text, truncated)
                    logger.debug(f"✅ Truncated cached embedding {len(cached_embedding)}→{target_dim}")
                    return truncated
                else:
                    # Dimension too small — discard, regenerate
                    logger.warning(f"Cached embedding dim {len(cached_embedding)} < {target_dim}, regenerating")

            # Cache miss - generate new embedding
            embedding = await self.embedding_manager.generate_embedding(text)

            # Cache for future use
            cache.set_embedding(text, embedding)
            logger.debug(f"💾 Cached new embedding")

            return embedding
        except Exception as e:
            logger.error(f"Error generating embedding: {e}")
            raise

    async def _generate_embeddings_batch(self, texts: List[str], max_concurrent: int = 5) -> List[List[float]]:
        """
        Generate embeddings for multiple texts with caching and parallel processing.

        Checks cache first, then batch-embeds only the cache misses.
        Uses parallel API calls for providers that support it (OpenRouter).
        """
        try:
            from core.cache import get_cache_service
            cache = get_cache_service()

            # Check cache for all texts
            cached = cache.get_embeddings_batch(texts)

            # Separate hits from misses (with dimension validation)
            miss_indices = []
            miss_texts = []
            results = [None] * len(texts)
            target_dim = self._get_target_dimension()
            stale_count = 0
            recache_entries = {}

            for i, text in enumerate(texts):
                emb = cached.get(text)
                if emb is not None:
                    if len(emb) == target_dim:
                        results[i] = emb
                    elif len(emb) > target_dim:
                        # Matryoshka: truncate stale cache entry
                        truncated = emb[:target_dim]
                        results[i] = truncated
                        recache_entries[text] = truncated
                        stale_count += 1
                    else:
                        # Too small — treat as miss
                        miss_indices.append(i)
                        miss_texts.append(text)
                        stale_count += 1
                else:
                    miss_indices.append(i)
                    miss_texts.append(text)

            cache_hits = len(texts) - len(miss_texts)
            if stale_count > 0:
                logger.warning(f"Embedding cache: {stale_count} stale entries (dim != {target_dim}), truncating/regenerating")
            if cache_hits > 0:
                logger.info(f"Embedding cache: {cache_hits}/{len(texts)} hits, {len(miss_texts)} misses")

            # Re-cache truncated entries
            if recache_entries:
                cache.set_embeddings_batch(recache_entries)
                logger.info(f"Re-cached {len(recache_entries)} truncated embeddings")

            if not miss_texts:
                return results

            # Batch embed the misses
            new_embeddings = await self.embedding_manager.generate_embeddings_batch(
                miss_texts, max_concurrent=max_concurrent
            )

            # Fill results and cache the new embeddings
            new_cache_entries = {}
            for idx, emb in zip(miss_indices, new_embeddings):
                results[idx] = emb
                new_cache_entries[texts[idx]] = emb

            if new_cache_entries:
                cache.set_embeddings_batch(new_cache_entries)
                logger.info(f"Cached {len(new_cache_entries)} new embeddings")

            return results

        except Exception as e:
            logger.warning(f"Batch embedding with cache failed ({e}), falling back to sequential")
            return [await self._generate_embedding(t) for t in texts]
    
    def list_documents(self, status: Optional[DocumentStatus] = None, 
                      file_type: Optional[DocumentType] = None,
                      limit: int = 100, offset: int = 0) -> List[Dict]:
        """List documents with optional filtering"""
        self._ensure_database_initialized()
        try:
            conn = psycopg2.connect(**self.db_config)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            query = "SELECT * FROM documents WHERE 1=1"
            params = []
            
            if status:
                query += " AND status = %s"
                params.append(status.value)
            
            if file_type:
                query += " AND file_type = %s"
                params.append(file_type.value)
            
            query += " ORDER BY upload_date DESC LIMIT %s OFFSET %s"
            params.extend([limit, offset])
            
            cursor.execute(query, params)
            documents = cursor.fetchall()
            
            cursor.close()
            conn.close()
            
            return [dict(doc) for doc in documents]
            
        except Exception as e:
            logger.error(f"Error listing documents: {e}")
            raise
    
    def get_document(self, document_id: int) -> Optional[Dict]:
        """Get document by ID"""
        self._ensure_database_initialized()
        try:
            conn = psycopg2.connect(**self.db_config)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            cursor.execute("SELECT * FROM documents WHERE id = %s", (document_id,))
            document = cursor.fetchone()
            
            cursor.close()
            conn.close()
            
            return dict(document) if document else None
            
        except Exception as e:
            logger.error(f"Error getting document {document_id}: {e}")
            raise
    
    def delete_document(self, document_id: int) -> bool:
        """Delete document, all its chunks, and any S3-stored vectors.

        §H 'delete removes the vector': in S3 Vectors mode the manager must
        also drop the chunk-vectors for the doc — Postgres-only delete left
        orphan vectors in the per-workspace index, breaking the contract.
        S3 cleanup is best-effort after the Postgres commit: a flaky S3
        outage MUST NOT block a delete the foreground row was already gone
        for. Logged loudly so the orphan is reconcilable; the next vector
        search filters by workspace + the doc_id is gone from Postgres so
        the orphan never surfaces.
        """
        self._ensure_database_initialized()
        try:
            conn = psycopg2.connect(**self.db_config)
            cursor = conn.cursor()

            # Delete chunks first (due to foreign key constraint)
            cursor.execute("DELETE FROM document_chunks WHERE document_id = %s", (document_id,))

            # Delete document
            cursor.execute("DELETE FROM documents WHERE id = %s", (document_id,))

            deleted = cursor.rowcount > 0
            conn.commit()
            cursor.close()
            conn.close()

            # S3 vector cleanup — only meaningful in S3 Vectors mode (legacy
            # pgvector path stores the vectors next to the chunks, already gone).
            if self.use_s3_vectors and self._s3_backend:
                try:
                    vectors_deleted = self._s3_backend.delete_documents(
                        str(document_id)
                    )
                    logger.info(
                        f"Deleted {vectors_deleted} S3 vectors for document "
                        f"{document_id}"
                    )
                except Exception:
                    logger.error(
                        f"S3 vector delete failed for doc {document_id} "
                        f"(postgres delete already committed)",
                        exc_info=True,
                    )

            logger.info(f"Document {document_id} deleted: {deleted}")
            return deleted

        except Exception as e:
            logger.error(f"Error deleting document {document_id}: {e}")
            raise
    
    def search_documents(self, query: str, limit: int = 10) -> List[Dict]:
        """Search documents by content similarity"""
        self._ensure_database_initialized()
        try:
            # Generate query embedding
            query_embedding = asyncio.run(self._generate_embedding(query))
            
            conn = psycopg2.connect(**self.db_config)
            cursor = conn.cursor(cursor_factory=RealDictCursor)
            
            # Format embedding for pgvector
            embedding_str = '[' + ','.join(map(str, query_embedding)) + ']'
            
            # Search for similar chunks using real vector similarity
            cursor.execute("""
                SELECT 
                    dc.id as chunk_id,
                    dc.document_id,
                    dc.content,
                    dc.metadata,
                    d.filename,
                    d.file_type,
                    1.0 as similarity
                FROM document_chunks dc
                JOIN documents d ON dc.document_id = d.id
                WHERE d.status = 'completed'
                    AND dc.embedding IS NOT NULL
                ORDER BY dc.embedding <=> %s::vector
                LIMIT %s
            """, (embedding_str, embedding_str, limit))
            
            results = cursor.fetchall()
            
            cursor.close()
            conn.close()
            
            return [dict(result) for result in results]
            
        except Exception as e:
            logger.error(f"Error searching documents: {e}")
            raise
    
    def get_document_stats(self) -> Dict:
        """Get document statistics"""
        self._ensure_database_initialized()
        try:
            conn = psycopg2.connect(**self.db_config)
            cursor = conn.cursor()
            
            # Total documents by status
            cursor.execute("""
                SELECT status, COUNT(*) as count
                FROM documents
                GROUP BY status
            """)
            status_counts = dict(cursor.fetchall())
            
            # Total chunks
            cursor.execute("SELECT COUNT(*) FROM document_chunks")
            total_chunks = cursor.fetchone()[0]
            
            # File type distribution
            cursor.execute("""
                SELECT file_type, COUNT(*) as count
                FROM documents
                GROUP BY file_type
            """)
            file_type_counts = dict(cursor.fetchall())
            
            cursor.close()
            conn.close()
            
            return {
                'total_documents': sum(status_counts.values()),
                'status_distribution': status_counts,
                'total_chunks': total_chunks,
                'file_type_distribution': file_type_counts
            }
            
        except Exception as e:
            logger.error(f"Error getting document stats: {e}")
            raise

# Example usage
if __name__ == "__main__":
    import asyncio
    from config import config

    # Database configuration
    db_config = {
        'host': config.POSTGRES_HOST or '127.0.0.1',
        'database': 'orchestrator_db',
        'user': 'postgres',
        'password': 'your_password'
    }
    
    # Initialize document manager
    doc_manager = DocumentManager(db_config, "your_openai_api_key")
    
    # Upload a document
    async def test_upload():
        doc_id = await doc_manager.upload_document(
            "/path/to/document.pdf",
            tags=["business", "process"],
            description="Business process documentation"
        )
        print(f"Uploaded document with ID: {doc_id}")
    
    # Run test
    # asyncio.run(test_upload())
